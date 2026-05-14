"""
RiceStock - Quintal System
FPS Rice Inventory Management System for Dealer: Balaram Shial
Developed by: Sunil Kumar
Version: 1.0.0
"""

import streamlit as st
import sqlite3
import pandas as pd
import hashlib
import os
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo
import calendar
from io import BytesIO

# ─────────────────────────────────────────────
# OPENPYXL IMPORTS
# ─────────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Alignment, PatternFill, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ─────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────
IST = ZoneInfo("Asia/Kolkata")
DB_PATH = "rice_inventory.db"
CARD_TYPES = ["PHH", "AAY", "SFSS"]
APP_TITLE = "RiceStock - Quintal System"
DEALER_NAME = "Dealer - Balaram Shial"
DEALER_CODE = "Code - 0201P100"
DEFAULT_USERNAME = "Sunil"
DEFAULT_PASSWORD = "sunil123"

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title=APP_TITLE,
    page_icon="🌾",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
# GLOBAL CSS THEME — Agricultural Green
# ─────────────────────────────────────────────
def inject_css():
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700;800&family=Playfair+Display:wght@700&display=swap');

    :root {
        --primary:        #1a6b3c;
        --primary-light:  #2e8b57;
        --primary-dark:   #0d4a28;
        --accent:         #f5a623;
        --accent-light:   #ffc55a;
        --bg-main:        #f4f9f4;
        --bg-card:        #ffffff;
        --bg-sidebar:     #0d4a28;
        --text-main:      #1c2e1c;
        --text-muted:     #5a7a5a;
        --text-light:     #ffffff;
        --border:         #c8e6c9;
        --success:        #388e3c;
        --warning:        #f57c00;
        --danger:         #c62828;
        --info:           #0277bd;
        --shadow:         0 4px 20px rgba(26,107,60,0.12);
        --radius:         12px;
    }

    html, body, [class*="css"] {
        font-family: 'Poppins', sans-serif;
        background-color: var(--bg-main);
        color: var(--text-main);
    }

    /* ── Sidebar ── */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0d4a28 0%, #1a6b3c 60%, #2e8b57 100%) !important;
        border-right: none !important;
    }
    [data-testid="stSidebar"] * { color: #e8f5e9 !important; }
    [data-testid="stSidebar"] .stRadio label { color: #e8f5e9 !important; font-size: 0.95rem; }
    [data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.2) !important; }
    [data-testid="stSidebarNav"] { display: none; }

    /* ── Main area ── */
    .main .block-container { padding-top: 1.5rem; padding-bottom: 2rem; max-width: 1200px; }

    /* ── Metric cards ── */
    [data-testid="stMetric"] {
        background: var(--bg-card);
        border: 1.5px solid var(--border);
        border-radius: var(--radius);
        padding: 1rem 1.2rem;
        box-shadow: var(--shadow);
        transition: transform 0.2s;
    }
    [data-testid="stMetric"]:hover { transform: translateY(-2px); }
    [data-testid="stMetricLabel"] { font-size: 0.8rem; color: var(--text-muted) !important; font-weight: 500; }
    [data-testid="stMetricValue"] { color: var(--primary) !important; font-weight: 700; font-size: 1.4rem !important; }

    /* ── Buttons ── */
    .stButton > button {
        background: linear-gradient(135deg, var(--primary) 0%, var(--primary-light) 100%);
        color: white !important;
        border: none;
        border-radius: 8px;
        padding: 0.55rem 1.8rem;
        font-weight: 600;
        font-size: 0.95rem;
        transition: all 0.25s ease;
        box-shadow: 0 3px 10px rgba(26,107,60,0.3);
        letter-spacing: 0.3px;
    }
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 18px rgba(26,107,60,0.4);
        background: linear-gradient(135deg, var(--primary-light) 0%, #3ea06a 100%);
    }

    /* ── Input fields ── */
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input,
    .stSelectbox > div > div,
    .stDateInput > div > div > input {
        border: 1.5px solid var(--border) !important;
        border-radius: 8px !important;
        background: #fff !important;
        color: var(--text-main) !important;
    }
    .stTextInput > div > div > input:focus,
    .stNumberInput > div > div > input:focus {
        border-color: var(--primary) !important;
        box-shadow: 0 0 0 3px rgba(26,107,60,0.12) !important;
    }

    /* ── Disabled / greyed fields ── */
    .stTextInput > div > div > input:disabled,
    .stNumberInput > div > div > input:disabled {
        background: #f0f4f0 !important;
        color: #8a9a8a !important;
        border-color: #d0ddd0 !important;
        cursor: not-allowed;
    }

    /* ── Custom cards ── */
    .card {
        background: var(--bg-card);
        border: 1.5px solid var(--border);
        border-radius: var(--radius);
        padding: 1.4rem 1.6rem;
        box-shadow: var(--shadow);
        margin-bottom: 1rem;
    }
    .card-header {
        font-size: 1.1rem;
        font-weight: 700;
        color: var(--primary);
        margin-bottom: 0.8rem;
        display: flex;
        align-items: center;
        gap: 0.5rem;
        border-bottom: 2px solid var(--border);
        padding-bottom: 0.6rem;
    }

    /* ── Page title ── */
    .page-title {
        font-family: 'Playfair Display', serif;
        font-size: 2rem;
        font-weight: 700;
        color: var(--primary-dark);
        margin-bottom: 0.2rem;
        line-height: 1.2;
    }
    .page-subtitle {
        font-size: 0.9rem;
        color: var(--text-muted);
        margin-bottom: 1.5rem;
    }

    /* ── Greeting banner ── */
    .greeting-banner {
        background: linear-gradient(135deg, var(--primary-dark) 0%, var(--primary) 50%, var(--primary-light) 100%);
        color: white;
        padding: 0.9rem 1.5rem;
        border-radius: var(--radius);
        margin-bottom: 1.4rem;
        display: flex;
        justify-content: space-between;
        align-items: center;
        box-shadow: 0 4px 15px rgba(13,74,40,0.35);
    }
    .greeting-text { font-size: 1.05rem; font-weight: 600; }
    .greeting-time { font-size: 0.82rem; opacity: 0.85; }

    /* ── Stock card ── */
    .stock-card {
        background: var(--bg-card);
        border-left: 5px solid var(--primary);
        border-radius: var(--radius);
        padding: 1.2rem 1.4rem;
        box-shadow: var(--shadow);
        margin-bottom: 0.8rem;
        transition: transform 0.2s;
    }
    .stock-card:hover { transform: translateX(4px); }
    .stock-card-title { font-size: 1rem; font-weight: 700; color: var(--primary); }
    .stock-card-value { font-size: 1.5rem; font-weight: 800; color: var(--text-main); margin: 0.3rem 0; }
    .stock-card-sub { font-size: 0.8rem; color: var(--text-muted); }

    /* ── Alert-style boxes ── */
    .info-box {
        background: #e3f2fd;
        border-left: 4px solid #1976d2;
        border-radius: 8px;
        padding: 0.8rem 1rem;
        margin: 0.6rem 0;
        font-size: 0.9rem;
        color: #0d47a1;
    }
    .success-box {
        background: #e8f5e9;
        border-left: 4px solid #2e7d32;
        border-radius: 8px;
        padding: 0.8rem 1rem;
        margin: 0.6rem 0;
        font-size: 0.9rem;
        color: #1b5e20;
    }
    .warning-box {
        background: #fff3e0;
        border-left: 4px solid #ef6c00;
        border-radius: 8px;
        padding: 0.8rem 1rem;
        margin: 0.6rem 0;
        font-size: 0.9rem;
        color: #bf360c;
    }

    /* ── Login page ── */
    .login-container {
        max-width: 420px;
        margin: 0 auto;
        background: white;
        border-radius: 20px;
        padding: 2.5rem 2.8rem;
        box-shadow: 0 20px 60px rgba(13,74,40,0.18);
        border-top: 5px solid var(--primary);
    }
    .login-logo {
        text-align: center;
        margin-bottom: 1.5rem;
    }
    .login-logo .emoji { font-size: 3.5rem; }
    .login-title {
        font-family: 'Playfair Display', serif;
        font-size: 1.5rem;
        font-weight: 700;
        color: var(--primary-dark);
        text-align: center;
        margin-bottom: 0.2rem;
    }
    .login-subtitle { font-size: 0.85rem; color: var(--text-muted); text-align: center; margin-bottom: 1.8rem; }
    .login-badge {
        background: var(--bg-main);
        border: 1px solid var(--border);
        border-radius: 8px;
        padding: 0.5rem 1rem;
        text-align: center;
        font-size: 0.82rem;
        color: var(--text-muted);
        margin-bottom: 1.5rem;
    }

    /* ── Footer ── */
    .footer {
        background: var(--primary-dark);
        color: #a5d6a7;
        text-align: center;
        padding: 1rem 1.5rem;
        border-radius: var(--radius);
        margin-top: 2.5rem;
        font-size: 0.85rem;
    }
    .footer strong { color: #e8f5e9; }

    /* ── Sidebar nav items ── */
    .nav-item {
        padding: 0.55rem 0.8rem;
        border-radius: 8px;
        margin: 0.15rem 0;
        cursor: pointer;
        font-size: 0.93rem;
        font-weight: 500;
        transition: background 0.2s;
    }
    .nav-item:hover { background: rgba(255,255,255,0.15); }
    .nav-item.active { background: rgba(255,255,255,0.25); font-weight: 700; }

    /* ── Table ── */
    .dataframe { font-size: 0.88rem !important; }
    thead th { background: var(--primary) !important; color: white !important; font-weight: 600 !important; }

    /* ── Progress bar override ── */
    .stProgress > div > div { background: var(--primary) !important; }

    /* ── Divider ── */
    hr { border-color: var(--border) !important; margin: 1.2rem 0; }

    /* ── Tabs ── */
    .stTabs [data-baseweb="tab"] {
        font-weight: 600;
        color: var(--text-muted);
    }
    .stTabs [aria-selected="true"] {
        color: var(--primary) !important;
        border-bottom-color: var(--primary) !important;
    }

    /* ── Section label ── */
    label { font-weight: 500 !important; font-size: 0.9rem !important; color: var(--text-main) !important; }

    /* ── Selectbox ── */
    [data-baseweb="select"] > div { border-radius: 8px !important; border-color: var(--border) !important; }

    /* ── Scrollbar ── */
    ::-webkit-scrollbar { width: 7px; height: 7px; }
    ::-webkit-scrollbar-track { background: #f0f4f0; }
    ::-webkit-scrollbar-thumb { background: var(--primary-light); border-radius: 4px; }
    </style>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────
# DATABASE LAYER
# ─────────────────────────────────────────────
def get_db_connection():
    """Return a SQLite connection with row_factory enabled."""
    conn = sqlite3.connect(DB_PATH, detect_types=sqlite3.PARSE_DECLTYPES)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """Initialise all tables and seed default credentials if needed."""
    conn = get_db_connection()
    c = conn.cursor()

    # Inventory entries table
    c.execute("""
        CREATE TABLE IF NOT EXISTS entries (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            entry_date      TEXT    NOT NULL,
            card_type       TEXT    NOT NULL,
            opening_bal     REAL    DEFAULT 0,
            received        REAL    DEFAULT 0,
            total           REAL    DEFAULT 0,
            closing_bal     REAL    DEFAULT 0,
            selling         REAL    DEFAULT 0,
            remarks         TEXT,
            created_at      TEXT    DEFAULT (datetime('now','localtime')),
            UNIQUE(entry_date, card_type)
        )
    """)

    # Credentials table
    c.execute("""
        CREATE TABLE IF NOT EXISTS credentials (
            id       INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT    NOT NULL UNIQUE,
            password TEXT    NOT NULL
        )
    """)

    # Seed default credentials if table is empty
    c.execute("SELECT COUNT(*) FROM credentials")
    if c.fetchone()[0] == 0:
        hashed = hash_password(DEFAULT_PASSWORD)
        c.execute(
            "INSERT INTO credentials (username, password) VALUES (?, ?)",
            (DEFAULT_USERNAME, hashed)
        )

    conn.commit()
    conn.close()


# ─────────────────────────────────────────────
# AUTH HELPERS
# ─────────────────────────────────────────────
def hash_password(plain: str) -> str:
    """SHA-256 hash of plain-text password."""
    return hashlib.sha256(plain.encode()).hexdigest()


def verify_login(username: str, password: str) -> bool:
    """Return True if credentials match the database record."""
    conn = get_db_connection()
    row = conn.execute(
        "SELECT password FROM credentials WHERE username = ? COLLATE NOCASE",
        (username,)
    ).fetchone()
    conn.close()
    if row and row["password"] == hash_password(password):
        return True
    return False


def get_credentials() -> dict:
    """Return the current username from DB."""
    conn = get_db_connection()
    row = conn.execute("SELECT username FROM credentials LIMIT 1").fetchone()
    conn.close()
    return {"username": row["username"]} if row else {"username": DEFAULT_USERNAME}


def update_credentials(new_username: str, new_password: str) -> bool:
    """Update username and/or password in the database."""
    try:
        conn = get_db_connection()
        conn.execute(
            "UPDATE credentials SET username = ?, password = ?",
            (new_username, hash_password(new_password))
        )
        conn.commit()
        conn.close()
        return True
    except Exception:
        return False


# ─────────────────────────────────────────────
# INVENTORY HELPERS
# ─────────────────────────────────────────────
def get_opening_balance(entry_date: date, card_type: str) -> float:
    """
    Fetch previous day's closing balance for the given card_type.
    Returns 0.0 if no prior record exists.
    """
    conn = get_db_connection()
    row = conn.execute(
        """
        SELECT closing_bal FROM entries
        WHERE entry_date < ? AND card_type = ?
        ORDER BY entry_date DESC LIMIT 1
        """,
        (entry_date.isoformat(), card_type)
    ).fetchone()
    conn.close()
    return float(row["closing_bal"]) if row else 0.0


def save_entry(
    entry_date: date,
    card_type: str,
    opening_bal: float,
    received: float,
    total: float,
    closing_bal: float,
    selling: float,
    remarks: str,
) -> bool:
    """
    Insert or replace an inventory entry.
    Returns True on success.
    """
    try:
        conn = get_db_connection()
        conn.execute(
            """
            INSERT INTO entries
                (entry_date, card_type, opening_bal, received, total,
                 closing_bal, selling, remarks)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(entry_date, card_type) DO UPDATE SET
                opening_bal = excluded.opening_bal,
                received    = excluded.received,
                total       = excluded.total,
                closing_bal = excluded.closing_bal,
                selling     = excluded.selling,
                remarks     = excluded.remarks,
                created_at  = datetime('now','localtime')
            """,
            (
                entry_date.isoformat(), card_type,
                opening_bal, received, total,
                closing_bal, selling, remarks
            )
        )
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        st.error(f"Database error: {e}")
        return False


def get_current_stock() -> dict:
    """Return the latest closing balance for each card type."""
    conn = get_db_connection()
    stock = {}
    for ct in CARD_TYPES:
        row = conn.execute(
            """
            SELECT closing_bal FROM entries
            WHERE card_type = ?
            ORDER BY entry_date DESC LIMIT 1
            """,
            (ct,)
        ).fetchone()
        stock[ct] = float(row["closing_bal"]) if row else 0.0
    conn.close()
    return stock


def get_monthly_data(year: int, month: int) -> pd.DataFrame:
    """Fetch all entries for a given year-month, ordered by date."""
    conn = get_db_connection()
    prefix = f"{year:04d}-{month:02d}-%"
    df = pd.read_sql_query(
        """
        SELECT entry_date, card_type, opening_bal, received, total,
               closing_bal, selling, remarks
        FROM entries
        WHERE entry_date LIKE ?
        ORDER BY entry_date, card_type
        """,
        conn, params=(prefix,)
    )
    conn.close()
    return df


def get_date_range_data(start_date: date, end_date: date) -> pd.DataFrame:
    """Fetch entries between two dates (inclusive)."""
    conn = get_db_connection()
    df = pd.read_sql_query(
        """
        SELECT entry_date, card_type, opening_bal, received, total,
               closing_bal, selling, remarks
        FROM entries
        WHERE entry_date BETWEEN ? AND ?
        ORDER BY entry_date, card_type
        """,
        conn,
        params=(start_date.isoformat(), end_date.isoformat())
    )
    conn.close()
    return df


def get_yearly_data(year: int) -> pd.DataFrame:
    """Fetch all entries for a calendar year."""
    conn = get_db_connection()
    df = pd.read_sql_query(
        """
        SELECT entry_date, card_type, opening_bal, received, total,
               closing_bal, selling, remarks
        FROM entries
        WHERE entry_date LIKE ?
        ORDER BY entry_date, card_type
        """,
        conn, params=(f"{year}-%",)
    )
    conn.close()
    return df


def get_30day_data() -> pd.DataFrame:
    """Fetch last 30 days of entries for the movement chart."""
    conn = get_db_connection()
    since = (datetime.now(IST) - timedelta(days=29)).date().isoformat()
    df = pd.read_sql_query(
        """
        SELECT entry_date, card_type, received, selling, closing_bal
        FROM entries
        WHERE entry_date >= ?
        ORDER BY entry_date
        """,
        conn, params=(since,)
    )
    conn.close()
    return df


def get_summary_stats() -> dict:
    """Return aggregate totals across all time."""
    conn = get_db_connection()
    row = conn.execute(
        "SELECT SUM(received) AS total_received, SUM(selling) AS total_sold FROM entries"
    ).fetchone()
    conn.close()
    stock = get_current_stock()
    return {
        "total_received": float(row["total_received"] or 0),
        "total_sold":     float(row["total_sold"] or 0),
        "current_stock":  sum(stock.values()),
    }


# ─────────────────────────────────────────────
# DISPLAY FORMAT HELPERS
# ─────────────────────────────────────────────
def kg_to_quintal_display(kg: float) -> str:
    """
    Convert Kg to human-readable Quintal notation:
      ≥ 100 Kg → 'X Q. YY Kg'
      < 100 Kg → 'YY Kg'
    """
    if kg < 0:
        kg = 0.0
    if kg >= 100:
        quintals   = int(kg // 100)
        remainder  = kg % 100
        if remainder == 0:
            return f"{quintals} Q."
        return f"{quintals} Q. {remainder:.0f} Kg"
    return f"{kg:.0f} Kg"


def kg_to_quintal_excel(kg: float) -> str:
    """Same as above but tighter format for Excel cells."""
    return kg_to_quintal_display(kg)


# ─────────────────────────────────────────────
# IST TIME HELPERS
# ─────────────────────────────────────────────
def now_ist() -> datetime:
    return datetime.now(IST)


def today_ist() -> date:
    return now_ist().date()


def get_greeting() -> str:
    hour = now_ist().hour
    if 5 <= hour < 12:
        return "Good Morning"
    elif 12 <= hour < 17:
        return "Good Afternoon"
    else:
        return "Good Evening"


def render_greeting_banner(username: str):
    """Render the dynamic IST greeting banner."""
    greeting  = get_greeting()
    now       = now_ist()
    time_str  = now.strftime("%A, %d %B %Y  |  %I:%M %p IST")
    st.markdown(
        f"""
        <div class="greeting-banner">
            <div>
                <div class="greeting-text">🌾 {greeting}, Mr. {username}!</div>
                <div class="greeting-time">Welcome back to RiceStock — Quintal System</div>
            </div>
            <div class="greeting-time">🕐 {time_str}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


# ─────────────────────────────────────────────
# FOOTER
# ─────────────────────────────────────────────
def render_footer():
    st.markdown(
        """
        <div class="footer">
            🌾 <strong>Rice Inventory</strong> &nbsp;|&nbsp;
            FPS Stock Register — Dealer: <strong>Balaram Shial</strong> &nbsp;|&nbsp;
            Created by — <strong>Sunil Kumar</strong>
        </div>
        """,
        unsafe_allow_html=True,
    )


# ─────────────────────────────────────────────
# SIDEBAR NAVIGATION
# ─────────────────────────────────────────────
def render_sidebar(username: str):
    with st.sidebar:
        st.markdown(
            """
            <div style="text-align:center;padding:1rem 0 0.5rem;">
                <div style="font-size:2.8rem;">🌾</div>
                <div style="font-size:1.15rem;font-weight:800;color:#e8f5e9;letter-spacing:0.5px;">
                    RiceStock
                </div>
                <div style="font-size:0.78rem;color:#a5d6a7;margin-top:0.2rem;">
                    Quintal System
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.markdown("---")

        # Dealer info
        st.markdown(
            f"""
            <div style="background:rgba(255,255,255,0.1);border-radius:10px;
                        padding:0.7rem 1rem;margin-bottom:0.8rem;">
                <div style="font-size:0.78rem;color:#81c784;">DEALER</div>
                <div style="font-size:0.9rem;font-weight:700;color:#e8f5e9;">Balaram Shial</div>
                <div style="font-size:0.75rem;color:#a5d6a7;">Code: 0201P100</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        nav_options = [
            ("🏠", "Dashboard"),
            ("📝", "Daily Entry"),
            ("📊", "Reports"),
            ("⚙️", "Settings"),
        ]
        choices = [f"{icon}  {label}" for icon, label in nav_options]
        selected = st.radio("Navigation", choices, label_visibility="collapsed")
        page = selected.split("  ", 1)[1].strip()

        st.markdown("---")
        # Session info
        st.markdown(
            f"""
            <div style="font-size:0.78rem;color:#a5d6a7;padding:0.3rem 0.2rem;">
                👤 Logged in as <strong style="color:#e8f5e9;">{username}</strong>
            </div>
            """,
            unsafe_allow_html=True,
        )

        if st.button("🚪  Logout", use_container_width=True):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()

    return page


# ─────────────────────────────────────────────
# PAGE: LOGIN
# ─────────────────────────────────────────────
def page_login():
    # Centre the login card
    _, col, _ = st.columns([1, 1.2, 1])
    with col:
        st.markdown(
            f"""
            <div class="login-container">
                <div class="login-logo"><span class="emoji">🌾</span></div>
                <div class="login-title">{APP_TITLE}</div>
                <div class="login-subtitle">
                    Government FPS Stock Management Portal
                </div>
                <div class="login-badge">
                    <strong>{DEALER_NAME}</strong><br>{DEALER_CODE}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        with st.form("login_form", clear_on_submit=False):
            st.markdown("#### 🔐 Sign In")
            username  = st.text_input("Username", placeholder="Enter username")
            password  = st.text_input("Password", type="password", placeholder="Enter password")
            submitted = st.form_submit_button("Login →", use_container_width=True)

        if submitted:
            if not username or not password:
                st.error("⚠️ Please enter both username and password.")
            elif verify_login(username, password):
                creds = get_credentials()
                st.session_state["authenticated"] = True
                st.session_state["username"]      = creds["username"]
                st.success("✅ Login successful! Redirecting…")
                st.rerun()
            else:
                st.error("❌ Invalid credentials. Please try again.")

        st.markdown(
            """
            <div style="text-align:center;font-size:0.8rem;color:#888;margin-top:1.2rem;">
                🌾 Rice Inventory &nbsp;|&nbsp; Created by Sunil Kumar
            </div>
            """,
            unsafe_allow_html=True,
        )


# ─────────────────────────────────────────────
# PAGE: DASHBOARD
# ─────────────────────────────────────────────
def page_dashboard(username: str):
    render_greeting_banner(username)

    st.markdown('<div class="page-title">📊 Dashboard</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-subtitle">Live stock overview and 30-day movement</div>', unsafe_allow_html=True)

    # ── Summary metrics ──
    stats = get_summary_stats()
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("📦 Total Received", kg_to_quintal_display(stats["total_received"]))
    with c2:
        st.metric("🛒 Total Sold", kg_to_quintal_display(stats["total_sold"]))
    with c3:
        st.metric("🏪 Current Stock", kg_to_quintal_display(stats["current_stock"]))
    with c4:
        total_entries = get_db_connection().execute("SELECT COUNT(*) FROM entries").fetchone()[0]
        st.metric("📝 Total Entries", str(total_entries))

    st.markdown("---")

    # ── Per-card stock ──
    stock = get_current_stock()
    max_stock = max(stock.values()) if any(stock.values()) else 1

    st.markdown("### 🌾 Current Stock by Card Type")
    col_phh, col_aay, col_sfss = st.columns(3)
    card_icons = {"PHH": "🟢", "AAY": "🔵", "SFSS": "🟠"}
    for ct, col in zip(CARD_TYPES, [col_phh, col_aay, col_sfss]):
        with col:
            val    = stock[ct]
            pct    = min(val / max(max_stock, 1), 1.0)
            colour = {"PHH": "#1a6b3c", "AAY": "#0277bd", "SFSS": "#ef6c00"}[ct]
            st.markdown(
                f"""
                <div class="stock-card" style="border-left-color:{colour};">
                    <div class="stock-card-title">{card_icons[ct]} {ct} Card</div>
                    <div class="stock-card-value">{kg_to_quintal_display(val)}</div>
                    <div class="stock-card-sub">{val:.0f} Kg total</div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            st.progress(pct)

    st.markdown("---")

    # ── 30-day movement chart ──
    st.markdown("### 📈 30-Day Stock Movement")
    df30 = get_30day_data()

    if df30.empty:
        st.markdown(
            """
            <div class="info-box">
                📭 No data available for the last 30 days. Start adding daily entries to see the chart.
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        # Aggregate by date
        chart_df = (
            df30.groupby("entry_date")
            .agg(received=("received", "sum"), selling=("selling", "sum"))
            .reset_index()
            .rename(columns={"entry_date": "Date", "received": "Received (Kg)", "selling": "Sold (Kg)"})
        )
        chart_df["Date"] = pd.to_datetime(chart_df["Date"])
        chart_df = chart_df.set_index("Date")
        st.line_chart(chart_df, color=["#1a6b3c", "#f5a623"], height=300)

    render_footer()


# ─────────────────────────────────────────────
# PAGE: DAILY ENTRY
# ─────────────────────────────────────────────
def page_daily_entry(username: str):
    render_greeting_banner(username)
    st.markdown('<div class="page-title">📝 Daily Entry</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-subtitle">Record daily rice stock transactions (all quantities in Kg)</div>', unsafe_allow_html=True)

    # ── Form ──
    with st.container():
        st.markdown('<div class="card"><div class="card-header">📋 New Transaction Entry</div>', unsafe_allow_html=True)

        col1, col2 = st.columns(2)
        with col1:
            entry_date = st.date_input(
                "📅 Date",
                value=today_ist(),
                max_value=today_ist(),
                help="Select transaction date (max: today)"
            )
        with col2:
            card_type = st.selectbox(
                "🃏 Card Type",
                CARD_TYPES,
                help="Select the ration card category"
            )

        # Auto-fetch opening balance
        opening_bal = get_opening_balance(entry_date, card_type)

        col3, col4 = st.columns(2)
        with col3:
            st.text_input(
                "📂 Opening Balance (auto-fetched)",
                value=kg_to_quintal_display(opening_bal),
                disabled=True,
                help="Automatically filled from previous day's closing balance"
            )
        with col4:
            received = st.number_input(
                "📥 Received (Kg)",
                min_value=0.0,
                step=1.0,
                value=0.0,
                format="%.0f",
                help="Enter quantity received today in Kg"
            )

        # Auto-calculated total
        total = opening_bal + received

        col5, col6 = st.columns(2)
        with col5:
            st.text_input(
                "🔢 Total (auto-calculated)",
                value=kg_to_quintal_display(total),
                disabled=True,
                help="Opening Balance + Received"
            )
        with col6:
            closing_bal = st.number_input(
                "📤 Closing Balance (Kg)",
                min_value=0.0,
                step=1.0,
                value=0.0,
                format="%.0f",
                help="Enter actual closing balance at end of day"
            )

        # ── Business logic: closing > total ──
        adjusted_received = received
        adjusted_closing  = closing_bal
        show_adjustment_msg = False

        if closing_bal > total:
            excess             = closing_bal - total
            adjusted_received  = received + excess
            adjusted_closing   = closing_bal            # keep what user entered
            total              = opening_bal + adjusted_received
            show_adjustment_msg = True

        selling = max(total - adjusted_closing, 0.0)

        # Show adjustment notice
        if show_adjustment_msg:
            st.markdown(
                f"""
                <div class="warning-box">
                    ⚠️ <strong>Auto-Adjusted:</strong> Closing Balance ({kg_to_quintal_display(closing_bal)})
                    exceeded Total ({kg_to_quintal_display(opening_bal + received)}).
                    Received has been automatically increased by {kg_to_quintal_display(excess)}
                    to balance the entry. New Received = {kg_to_quintal_display(adjusted_received)}.
                </div>
                """,
                unsafe_allow_html=True,
            )

        col7, col8 = st.columns(2)
        with col7:
            st.text_input(
                "💰 Selling / Issued (auto-calculated)",
                value=kg_to_quintal_display(selling),
                disabled=True,
                help="Total − Closing Balance"
            )
        with col8:
            remarks = st.text_input(
                "📌 Remarks (optional)",
                placeholder="Any notes for this entry…"
            )

        st.markdown("</div>", unsafe_allow_html=True)

        # ── Preview ──
        st.markdown("#### 📋 Entry Preview")
        preview_data = {
            "Field":  ["Date", "Card Type", "Opening Balance", "Received", "Total", "Closing Balance", "Selling", "Remarks"],
            "Value":  [
                entry_date.strftime("%d %B %Y"),
                card_type,
                kg_to_quintal_display(opening_bal),
                kg_to_quintal_display(adjusted_received),
                kg_to_quintal_display(total),
                kg_to_quintal_display(adjusted_closing),
                kg_to_quintal_display(selling),
                remarks or "—",
            ],
        }
        st.table(pd.DataFrame(preview_data))

        # ── Save button ──
        _, btn_col, _ = st.columns([2, 1, 2])
        with btn_col:
            if st.button("💾  Save Entry", use_container_width=True):
                ok = save_entry(
                    entry_date,
                    card_type,
                    opening_bal,
                    adjusted_received,
                    total,
                    adjusted_closing,
                    selling,
                    remarks,
                )
                if ok:
                    st.markdown(
                        f"""
                        <div class="success-box">
                            ✅ Entry saved successfully for <strong>{card_type}</strong>
                            on <strong>{entry_date.strftime('%d %B %Y')}</strong>.
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )
                else:
                    st.error("❌ Failed to save entry. Please try again.")

    render_footer()


# ─────────────────────────────────────────────
# EXCEL EXPORT HELPERS
# ─────────────────────────────────────────────
def build_excel_report(df: pd.DataFrame, title: str, header_label: str) -> BytesIO:
    """
    Build a government-style FPS Stock Register Excel file.
    Returns a BytesIO stream ready for download.
    """
    wb = Workbook()

    for ct in CARD_TYPES:
        sub = df[df["card_type"] == ct].copy()
        ws  = wb.create_sheet(title=ct)

        # ── Styles ──
        header_fill  = PatternFill("solid", fgColor="1A6B3C")
        subhdr_fill  = PatternFill("solid", fgColor="E8F5E9")
        title_font   = Font(name="Arial", bold=True, size=14, color="FFFFFF")
        col_font     = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        normal_font  = Font(name="Arial", size=10)
        bold_font    = Font(name="Arial", bold=True, size=10)
        centre       = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        thin         = Side(border_style="thin", color="AAAAAA")
        thick        = Side(border_style="medium", color="1A6B3C")
        thin_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
        thick_border = Border(left=thick, right=thick, top=thick, bottom=thick)

        # ── Row 1: Main title ──
        ws.merge_cells("A1:H1")
        ws["A1"] = "F.P.S. STOCK REGISTER"
        ws["A1"].font      = Font(name="Arial", bold=True, size=16, color="1A6B3C")
        ws["A1"].alignment = centre

        # ── Row 2: Sub-title ──
        ws.merge_cells("A2:H2")
        ws["A2"] = f"F.P.S. Code No. : {DEALER_CODE.replace('Code - ', '')}"
        ws["A2"].font      = Font(name="Arial", bold=True, size=11)
        ws["A2"].alignment = centre

        # ── Row 3: Dealer info ──
        ws.merge_cells("A3:H3")
        ws["A3"] = f"Dealer: {DEALER_NAME.replace('Dealer - ', '')}   |   Card Type: {ct}"
        ws["A3"].font      = Font(name="Arial", size=10)
        ws["A3"].alignment = centre

        # ── Row 4: Report header ──
        ws.merge_cells("A4:H4")
        ws["A4"] = header_label
        ws["A4"].font      = Font(name="Arial", bold=True, size=11)
        ws["A4"].alignment = centre
        ws["A4"].fill      = PatternFill("solid", fgColor="C8E6C9")

        # ── Row 5: Column headers ──
        headers = [
            "Date of\nTransaction",
            "Opening Balance\n(Quintal)",
            "Receipt\n(Quintal)",
            "Total\n(Quintal)",
            "No. of\nCards",
            "Quantity\n(Issued)",
            "Closing Balance\n(Quintal)",
            "Remarks",
        ]
        for col_idx, h in enumerate(headers, start=1):
            cell            = ws.cell(row=5, column=col_idx, value=h)
            cell.font       = col_font
            cell.fill       = header_fill
            cell.alignment  = centre
            cell.border     = thin_border

        # ── Column widths ──
        col_widths = [16, 18, 16, 16, 12, 16, 18, 22]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ── Data rows ──
        row_num = 6
        total_received   = 0.0
        total_selling    = 0.0
        total_closing    = 0.0

        if sub.empty:
            ws.merge_cells(f"A{row_num}:H{row_num}")
            ws.cell(row=row_num, column=1, value="No data for this period.").alignment = centre
            row_num += 1
        else:
            for _, r in sub.iterrows():
                data_row = [
                    datetime.strptime(r["entry_date"], "%Y-%m-%d").strftime("%d-%m-%Y"),
                    kg_to_quintal_excel(r["opening_bal"]),
                    kg_to_quintal_excel(r["received"]),
                    kg_to_quintal_excel(r["total"]),
                    "",                                  # No. of Cards – blank per spec
                    kg_to_quintal_excel(r["selling"]),
                    kg_to_quintal_excel(r["closing_bal"]),
                    r["remarks"] or "",
                ]
                fill = PatternFill("solid", fgColor="F9FFF9") if row_num % 2 == 0 else None
                for col_idx, val in enumerate(data_row, start=1):
                    cell           = ws.cell(row=row_num, column=col_idx, value=val)
                    cell.font      = normal_font
                    cell.alignment = centre if col_idx != 1 else left_align
                    cell.border    = thin_border
                    if fill:
                        cell.fill = fill

                total_received += float(r["received"])
                total_selling  += float(r["selling"])
                total_closing   = float(r["closing_bal"])   # last closing
                row_num += 1

            # ── Totals row ──
            ws.merge_cells(f"A{row_num}:A{row_num}")
            totals = [
                "TOTALS",
                "",
                kg_to_quintal_excel(total_received),
                "",
                "",
                kg_to_quintal_excel(total_selling),
                kg_to_quintal_excel(total_closing),
                "",
            ]
            total_fill = PatternFill("solid", fgColor="C8E6C9")
            for col_idx, val in enumerate(totals, start=1):
                cell           = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font      = bold_font
                cell.fill      = total_fill
                cell.alignment = centre
                cell.border    = thin_border

            # ── Signature row ──
            row_num += 2
            ws.merge_cells(f"A{row_num}:D{row_num}")
            ws.cell(row=row_num, column=1, value="Checked by / Inspector Signature:").font = bold_font
            ws.merge_cells(f"F{row_num}:H{row_num}")
            ws.cell(row=row_num, column=6, value="Dealer Signature: _____________").font = bold_font

        # ── Row heights ──
        ws.row_dimensions[5].height = 42
        for rn in range(1, row_num + 1):
            if rn < 5:
                ws.row_dimensions[rn].height = 22

    # Remove default empty sheet if needed
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── Combined sheet if multiple card types ──
    if not df.empty:
        ws_all = wb.create_sheet(title="COMBINED", index=0)
        header_fill_c = PatternFill("solid", fgColor="1A6B3C")
        col_font_c    = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        normal_font_c = Font(name="Arial", size=10)
        thin_c        = Side(border_style="thin", color="AAAAAA")
        thin_border_c = Border(left=thin_c, right=thin_c, top=thin_c, bottom=thin_c)
        centre_c      = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws_all.merge_cells("A1:I1")
        ws_all["A1"] = "F.P.S. STOCK REGISTER — COMBINED SUMMARY"
        ws_all["A1"].font      = Font(name="Arial", bold=True, size=14, color="1A6B3C")
        ws_all["A1"].alignment = centre_c

        ws_all.merge_cells("A2:I2")
        ws_all["A2"] = header_label
        ws_all["A2"].font      = Font(name="Arial", bold=True, size=11)
        ws_all["A2"].alignment = centre_c
        ws_all["A2"].fill      = PatternFill("solid", fgColor="C8E6C9")

        combo_headers = [
            "Date", "Card Type", "Opening Balance", "Receipt",
            "Total", "No. of Cards", "Qty Issued", "Closing Balance", "Remarks"
        ]
        for ci, h in enumerate(combo_headers, start=1):
            cell           = ws_all.cell(row=3, column=ci, value=h)
            cell.font      = col_font_c
            cell.fill      = header_fill_c
            cell.alignment = centre_c
            cell.border    = thin_border_c

        col_w = [16, 12, 18, 16, 16, 12, 16, 18, 22]
        for i, w in enumerate(col_w, start=1):
            ws_all.column_dimensions[get_column_letter(i)].width = w

        rn = 4
        for _, r in df.sort_values(["entry_date", "card_type"]).iterrows():
            row_data = [
                datetime.strptime(r["entry_date"], "%Y-%m-%d").strftime("%d-%m-%Y"),
                r["card_type"],
                kg_to_quintal_excel(r["opening_bal"]),
                kg_to_quintal_excel(r["received"]),
                kg_to_quintal_excel(r["total"]),
                "",
                kg_to_quintal_excel(r["selling"]),
                kg_to_quintal_excel(r["closing_bal"]),
                r["remarks"] or "",
            ]
            alt_fill = PatternFill("solid", fgColor="F9FFF9") if rn % 2 == 0 else None
            for ci, v in enumerate(row_data, start=1):
                cell           = ws_all.cell(row=rn, column=ci, value=v)
                cell.font      = normal_font_c
                cell.alignment = centre_c
                cell.border    = thin_border_c
                if alt_fill:
                    cell.fill = alt_fill
            rn += 1

        ws_all.row_dimensions[3].height = 40

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────
# PAGE: REPORTS
# ─────────────────────────────────────────────
def page_reports(username: str):
    render_greeting_banner(username)
    st.markdown('<div class="page-title">📊 Reports & Export</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-subtitle">View monthly/yearly summaries and export FPS Stock Register</div>', unsafe_allow_html=True)

    tabs = st.tabs(["📅 Monthly Report", "📆 Yearly Summary", "📤 Export"])

    # ── Tab 1: Monthly ──
    with tabs[0]:
        col1, col2 = st.columns(2)
        with col1:
            month = st.selectbox("Month", list(range(1, 13)),
                                 format_func=lambda m: calendar.month_name[m],
                                 index=today_ist().month - 1)
        with col2:
            year = st.number_input("Year", min_value=2020, max_value=today_ist().year + 1,
                                   value=today_ist().year, step=1)

        df = get_monthly_data(year, month)
        period_label = f"{calendar.month_name[month]} {year}"

        if df.empty:
            st.markdown(
                f'<div class="info-box">📭 No entries found for {period_label}.</div>',
                unsafe_allow_html=True,
            )
        else:
            # Display per card type
            for ct in CARD_TYPES:
                sub = df[df["card_type"] == ct]
                if sub.empty:
                    continue
                st.markdown(f"#### 🃏 {ct} — {period_label}")
                display_df = sub[["entry_date","opening_bal","received","total","closing_bal","selling","remarks"]].copy()
                for col_kg in ["opening_bal","received","total","closing_bal","selling"]:
                    display_df[col_kg] = display_df[col_kg].apply(kg_to_quintal_display)
                display_df.columns = ["Date","Opening","Received","Total","Closing","Selling","Remarks"]
                display_df["Date"] = pd.to_datetime(display_df["Date"]).dt.strftime("%d %b %Y")
                st.dataframe(display_df, use_container_width=True, hide_index=True)

    # ── Tab 2: Yearly ──
    with tabs[1]:
        yr = st.number_input("Select Year", min_value=2020, max_value=today_ist().year + 1,
                             value=today_ist().year, step=1, key="yr_sel")
        df_yr = get_yearly_data(yr)

        if df_yr.empty:
            st.markdown(
                f'<div class="info-box">📭 No entries found for {yr}.</div>',
                unsafe_allow_html=True,
            )
        else:
            # Monthly aggregation
            df_yr["month"] = pd.to_datetime(df_yr["entry_date"]).dt.month
            agg = (
                df_yr.groupby(["month","card_type"])
                .agg(received=("received","sum"), selling=("selling","sum"))
                .reset_index()
            )
            pivot = agg.pivot_table(
                index="month", columns="card_type",
                values=["received","selling"], aggfunc="sum", fill_value=0
            )
            pivot.index = [calendar.month_abbr[m] for m in pivot.index]
            st.dataframe(pivot.round(0).astype(int), use_container_width=True)

            st.markdown("##### 📊 Yearly Received vs Sold (Kg)")
            chart_data = df_yr.groupby("entry_date").agg(
                Received=("received","sum"), Sold=("selling","sum")
            )
            chart_data.index = pd.to_datetime(chart_data.index)
            st.bar_chart(chart_data, color=["#1a6b3c","#f5a623"])

    # ── Tab 3: Export ──
    with tabs[2]:
        if not OPENPYXL_AVAILABLE:
            st.error("❌ openpyxl library not installed. Please add it to requirements.txt.")
            return

        export_type = st.selectbox(
            "Export Type",
            ["Single Month", "Multiple Months (Combined)", "Full Year", "Custom Date Range"],
        )

        if export_type == "Single Month":
            c1, c2 = st.columns(2)
            with c1:
                em = st.selectbox("Month", list(range(1,13)),
                                  format_func=lambda x: calendar.month_name[x],
                                  index=today_ist().month - 1, key="exp_m")
            with c2:
                ey = st.number_input("Year", min_value=2020, max_value=today_ist().year + 1,
                                     value=today_ist().year, step=1, key="exp_y")
            if st.button("📥 Export Single Month"):
                df = get_monthly_data(ey, em)
                label = f"Month / Year — {calendar.month_name[em]} {ey}"
                buf   = build_excel_report(df, label, label)
                fname = f"FPS_Stock_{calendar.month_abbr[em]}_{ey}.xlsx"
                st.download_button("⬇️ Download Excel", buf, fname,
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        elif export_type == "Multiple Months (Combined)":
            st.markdown("Select 2 or more months:")
            myr = st.number_input("Year", min_value=2020, max_value=today_ist().year + 1,
                                  value=today_ist().year, step=1, key="mmyr")
            selected_months = st.multiselect(
                "Months",
                list(range(1, 13)),
                default=[today_ist().month],
                format_func=lambda x: calendar.month_name[x],
            )
            if st.button("📥 Export Combined Months"):
                if len(selected_months) < 2:
                    st.warning("⚠️ Please select at least 2 months.")
                else:
                    frames = [get_monthly_data(myr, m) for m in sorted(selected_months)]
                    df     = pd.concat(frames, ignore_index=True)
                    m_names = ", ".join(calendar.month_abbr[m] for m in sorted(selected_months))
                    label   = f"Combined Report — {m_names} {myr}"
                    buf     = build_excel_report(df, label, label)
                    fname   = f"FPS_Combined_{m_names}_{myr}.xlsx"
                    st.download_button("⬇️ Download Excel", buf, fname,
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        elif export_type == "Full Year":
            fy = st.number_input("Year", min_value=2020, max_value=today_ist().year + 1,
                                 value=today_ist().year, step=1, key="full_yr")
            if st.button("📥 Export Full Year"):
                df    = get_yearly_data(fy)
                label = f"Full Year Report — {fy}"
                buf   = build_excel_report(df, label, label)
                fname = f"FPS_FullYear_{fy}.xlsx"
                st.download_button("⬇️ Download Excel", buf, fname,
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        else:  # Custom Date Range
            c1, c2 = st.columns(2)
            with c1:
                start_d = st.date_input("Start Date", value=today_ist() - timedelta(days=30), key="cdr_s")
            with c2:
                end_d   = st.date_input("End Date", value=today_ist(), key="cdr_e")
            if st.button("📥 Export Custom Range"):
                if start_d > end_d:
                    st.error("❌ Start date must be before end date.")
                else:
                    df    = get_date_range_data(start_d, end_d)
                    label = f"Custom Range — {start_d.strftime('%d %b %Y')} to {end_d.strftime('%d %b %Y')}"
                    buf   = build_excel_report(df, label, label)
                    fname = f"FPS_Custom_{start_d}_{end_d}.xlsx"
                    st.download_button("⬇️ Download Excel", buf, fname,
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    render_footer()


# ─────────────────────────────────────────────
# PAGE: SETTINGS
# ─────────────────────────────────────────────
def page_settings(username: str):
    render_greeting_banner(username)
    st.markdown('<div class="page-title">⚙️ Settings</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-subtitle">Manage your login credentials</div>', unsafe_allow_html=True)

    with st.container():
        st.markdown('<div class="card"><div class="card-header">🔐 Change Credentials</div>', unsafe_allow_html=True)

        current_creds = get_credentials()
        st.markdown(
            f'<div class="info-box">ℹ️ Current username: <strong>{current_creds["username"]}</strong></div>',
            unsafe_allow_html=True,
        )

        with st.form("settings_form"):
            new_username = st.text_input(
                "New Username",
                value=current_creds["username"],
                placeholder="Enter new username"
            )
            current_pass = st.text_input("Current Password", type="password", placeholder="Verify current password")
            new_pass1    = st.text_input("New Password", type="password", placeholder="Enter new password")
            new_pass2    = st.text_input("Confirm New Password", type="password", placeholder="Re-enter new password")
            submitted    = st.form_submit_button("💾  Save Changes", use_container_width=True)

        if submitted:
            if not current_pass:
                st.error("⚠️ Please enter your current password to make changes.")
            elif not verify_login(current_creds["username"], current_pass):
                st.error("❌ Current password is incorrect.")
            elif not new_username.strip():
                st.error("⚠️ Username cannot be empty.")
            elif new_pass1 and new_pass1 != new_pass2:
                st.error("❌ New passwords do not match.")
            elif new_pass1 and len(new_pass1) < 6:
                st.error("⚠️ Password must be at least 6 characters long.")
            else:
                final_pass = new_pass1 if new_pass1 else current_pass
                ok = update_credentials(new_username.strip(), final_pass)
                if ok:
                    st.session_state["username"] = new_username.strip()
                    st.markdown(
                        '<div class="success-box">✅ Credentials updated successfully! Sidebar will reflect new username.</div>',
                        unsafe_allow_html=True,
                    )
                else:
                    st.error("❌ Failed to update credentials.")

        st.markdown("</div>", unsafe_allow_html=True)

    # ── About section ──
    st.markdown("---")
    st.markdown(
        """
        <div class="card">
            <div class="card-header">ℹ️ About RiceStock</div>
            <table style="width:100%;font-size:0.9rem;">
                <tr><td style="color:#5a7a5a;width:45%;">Application</td><td><strong>RiceStock — Quintal System</strong></td></tr>
                <tr><td style="color:#5a7a5a;">Version</td><td>1.0.0</td></tr>
                <tr><td style="color:#5a7a5a;">Dealer</td><td>Balaram Shial</td></tr>
                <tr><td style="color:#5a7a5a;">FPS Code</td><td>0201P100</td></tr>
                <tr><td style="color:#5a7a5a;">Database</td><td>SQLite (rice_inventory.db)</td></tr>
                <tr><td style="color:#5a7a5a;">Timezone</td><td>Indian Standard Time (IST, UTC+5:30)</td></tr>
                <tr><td style="color:#5a7a5a;">Developed by</td><td><strong>Sunil Kumar</strong></td></tr>
            </table>
        </div>
        """,
        unsafe_allow_html=True,
    )

    render_footer()


# ─────────────────────────────────────────────
# MAIN APP ENTRY POINT
# ─────────────────────────────────────────────
def main():
    # Initialise DB on every run (idempotent)
    init_db()

    # Inject global CSS
    inject_css()

    # ── Auth gate ──
    if not st.session_state.get("authenticated", False):
        page_login()
        return

    username = st.session_state.get("username", DEFAULT_USERNAME)

    # ── Sidebar + routing ──
    page = render_sidebar(username)

    if page == "Dashboard":
        page_dashboard(username)
    elif page == "Daily Entry":
        page_daily_entry(username)
    elif page == "Reports":
        page_reports(username)
    elif page == "Settings":
        page_settings(username)


if __name__ == "__main__":
    main()
